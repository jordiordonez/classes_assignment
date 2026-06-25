"""Génération des feuilles de sortie (Classes, Contraintes, Impossibilites,
Tableau, Dashboards) — partagé par l'affectation initiale et la mise à jour
manuelle, afin d'éviter toute duplication."""

import io
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from xlsxwriter.utility import xl_col_to_name


# Classeur modèle (.xlsm) dans lequel la macro « Desanonymiser » a été incrustée
# manuellement. On part de CE fichier (créé par Excel, donc valide) et on y écrit
# les données : la sortie hérite ainsi d'un projet VBA intact, ouvrable par Excel.
VBA_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "Anonymiser", "assignments_desanonymiser.xlsm"
)

# Ordre d'affichage des onglets dans le classeur de sortie.
SHEET_ORDER = ["Classes", "Impossibilites", "Contraintes", "Tableau", "Dashboards", "diccionari"]


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Place la colonne `classe` juste après `student`."""
    cols = df.columns.tolist()
    if "student" in cols and "classe" in cols:
        cols.remove("classe")
        cols.insert(cols.index("student") + 1, "classe")
        return df[cols]
    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Statistiques agrégées par classe (effectifs + pourcentages)."""
    df = df.copy()
    # Colonnes optionnelles : présentes selon le format du fichier d'entrée.
    for col in ("por", "lat", "tech", "cat", "pap"):
        if col not in df.columns:
            df[col] = 0
    summary = df.groupby("classe").agg(
        Total=("student", "count"),
        Niveau1=("level", lambda x: (x == 1).sum()),
        Niveau2=("level", lambda x: (x == 2).sum()),
        Niveau3=("level", lambda x: (x == 3).sum()),
        POR=("por", "sum"),
        LAT=("lat", "sum"),
        TECH=("tech", "sum"),
        CAT=("cat", "sum"),
        PAP=("pap", "sum"),
        Filles=("Genre", lambda x: (x == "F").sum()),
        Garçons=("Genre", lambda x: (x == "G").sum()),
        Comp1=("Comportement", lambda x: (x == 1).sum()),
        Comp2=("Comportement", lambda x: (x == 2).sum()),
        Comp3=("Comportement", lambda x: (x == 3).sum()),
    )
    pct = {
        "%N1": "Niveau1", "%N2": "Niveau2", "%N3": "Niveau3",
        "%Filles": "Filles", "%Garçons": "Garçons",
        "%C1": "Comp1", "%C2": "Comp2", "%C3": "Comp3",
        "%PAP": "PAP",
    }
    for out, src in pct.items():
        summary[out] = (summary[src] / summary["Total"] * 100).round(0)

    # Origine (provenance A–H) : nombre d'élèves de chaque origine par classe.
    if "Origine" in df.columns:
        orig = df.assign(Origine=df["Origine"].astype(str).str.strip())
        orig = orig[orig["Origine"].isin(list("ABCDEFGH"))]
        if not orig.empty:
            counts = pd.crosstab(orig["classe"], orig["Origine"])
            counts.columns = [f"Origine_{c}" for c in counts.columns]
            summary = summary.join(counts)
            for c in counts.columns:
                summary[c] = summary[c].fillna(0).astype(int)
    return summary


def compute_constraints(students_df: pd.DataFrame, assign: dict):
    """Construit les DataFrames Contraintes (tous les vœux) et Impossibilites
    (vœux non respectés), à partir des données brutes et d'une affectation.

    `students_df` doit contenir la colonne `student` (nom) et les champs
    avec1/avec2 et sans1…sans5. `assign` mappe un nom d'élève vers sa classe.
    """
    avec_fields = ("avec1", "avec2")
    sans_fields = ("sans1", "sans2", "sans3", "sans4", "sans5")
    constraints, impossibilities = [], []
    for _, row in students_df.iterrows():
        student = str(row["student"])
        for fld in avec_fields:
            other = row.get(fld)
            if pd.notna(other):
                constraints.append([fld, student, str(other)])
                if assign.get(student) != assign.get(str(other)):
                    impossibilities.append([fld, student, str(other)])
        for fld in sans_fields:
            other = row.get(fld)
            if pd.notna(other):
                constraints.append([fld, student, str(other)])
                if assign.get(student) == assign.get(str(other)):
                    impossibilities.append([fld, student, str(other)])

    cols = ["Type", "Source", "Other"]
    constraints_df = pd.DataFrame(constraints, columns=cols).sort_values(["Type", "Source"])
    impossibilites_df = pd.DataFrame(impossibilities, columns=cols).sort_values(["Type", "Source"])
    return constraints_df, impossibilites_df


def generate_tableau_sheet(writer, students_df: pd.DataFrame, summary: pd.DataFrame):
    """Feuille matricielle : élèves par classe + sparklines (genre/niveau/comp)."""
    students_df = students_df.sort_values(by=["classe", "student"])
    tableau_data = {cls: list(grp["student"]) for cls, grp in students_df.groupby("classe")}
    max_len = max((len(lst) for lst in tableau_data.values()), default=0)
    tableau_padded = {
        cls: lst + [""] * (max_len - len(lst)) for cls, lst in tableau_data.items()
    }
    tableau_df = pd.DataFrame(tableau_padded)

    tableau_df.to_excel(writer, sheet_name="Tableau", startrow=7, startcol=1, index=False)
    ws = writer.sheets["Tableau"]

    HDR, FG_LBL, FG_CH, LVL_LBL, LVL_CH, CMP_LBL, CMP_CH, PL_LBL = range(8)
    STU_START = 8
    base_row = 100

    def stat(cl, key):
        return summary.loc[cl, key] if cl in summary.index else 0

    classes = tableau_df.columns.tolist()
    for j, cl in enumerate(classes):
        col = j + 1
        col_letter = xl_col_to_name(col)
        ws.write(HDR, col, cl)
        ws.write(FG_LBL, col, "       F            G")
        ws.write(LVL_LBL, col, "   N1      N2     N3")
        ws.write(CMP_LBL, col, "   C1      C2     C3")
        ws.write(PL_LBL, col, (
            f"POR:{int(stat(cl, 'POR'))} LAT:{int(stat(cl, 'LAT'))} "
            f"TECH:{int(stat(cl, 'TECH'))} CAT:{int(stat(cl, 'CAT'))} "
            f"PAP:{int(stat(cl, 'PAP'))}"
        ))

        for offset, key in enumerate(
            ["Filles", "Garçons", "Niveau1", "Niveau2", "Niveau3", "Comp1", "Comp2", "Comp3"]
        ):
            ws.write(base_row + offset, col, stat(cl, key))

        ws.add_sparkline(FG_CH, col, {
            "range": f"{col_letter}{base_row + 1}:{col_letter}{base_row + 2}",
            "type": "column", "min": 0,
        })
        ws.add_sparkline(LVL_CH, col, {
            "range": f"{col_letter}{base_row + 3}:{col_letter}{base_row + 5}",
            "type": "column", "min": 0,
        })
        ws.add_sparkline(CMP_CH, col, {
            "range": f"{col_letter}{base_row + 6}:{col_letter}{base_row + 8}",
            "type": "column", "min": 0,
        })

    for i in range(len(tableau_df)):
        ws.write(STU_START + i, 0, i + 1)

    for r in (FG_LBL, LVL_LBL, CMP_LBL):
        ws.set_row(r, 15)
    for r in (FG_CH, LVL_CH, CMP_CH):
        ws.set_row(r, 30)
    ws.set_column(0, 0, 5)
    ws.set_column(1, len(classes), 15)
    return tableau_df


def build_workbook(students_sorted: pd.DataFrame, constraints_df: pd.DataFrame,
                   impossibilites_df: pd.DataFrame, summary: pd.DataFrame,
                   with_macros: bool = True) -> io.BytesIO:
    """Assemble le classeur de sortie et renvoie un buffer prêt à télécharger.

    Si `with_macros` et que le modèle `.xlsm` existe, on ÉCRIT dans une copie du
    modèle (en conservant son projet VBA via openpyxl) : la sortie est un `.xlsm`
    macro-actif et valide pour Excel. L'utilisateur colle ensuite l'onglet
    `diccionari` (nom_real / id_anonim) et lance la macro « Desanonymiser ».
    Sinon, on retombe sur un `.xlsx` classique généré par xlsxwriter.
    """
    students_sorted = reorder_columns(students_sorted).sort_values(["classe", "student"])

    if with_macros and os.path.exists(VBA_TEMPLATE):
        return _build_macro_workbook(
            students_sorted, constraints_df, impossibilites_df, summary
        )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        students_sorted.to_excel(writer, sheet_name="Classes", index=False)
        if not impossibilites_df.empty:
            impossibilites_df.to_excel(writer, sheet_name="Impossibilites", index=False)
        if not constraints_df.empty:
            constraints_df.to_excel(writer, sheet_name="Contraintes", index=False)
        generate_tableau_sheet(writer, students_sorted, summary)
        summary.to_excel(writer, sheet_name="Dashboards")
    buffer.seek(0)
    return buffer


# ──────────────────────────────────────────────────────────────────────────
#  Sortie macro-active : on remplit le modèle .xlsm sans casser son VBA
# ──────────────────────────────────────────────────────────────────────────
def _build_macro_workbook(students_sorted, constraints_df, impossibilites_df, summary):
    wb = load_workbook(VBA_TEMPLATE, keep_vba=True)

    _write_df_sheet(wb, "Classes", students_sorted)
    _write_df_sheet(wb, "Impossibilites", impossibilites_df)
    _write_df_sheet(wb, "Contraintes", constraints_df)
    _write_tableau_sheet(wb, students_sorted, summary)
    _write_df_sheet(wb, "Dashboards", summary.reset_index())
    _reset_dictionary_sheet(wb)

    # Réordonne les onglets (cosmétique ; le VBA cible les feuilles par nom).
    wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title)
                    if s.title in SHEET_ORDER else len(SHEET_ORDER))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _fresh_sheet(wb, name):
    """Repart d'une feuille vide : supprime l'éventuelle existante puis recrée."""
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_df_sheet(wb, name, df):
    """Écrit un DataFrame (en-têtes + données) dans une feuille fraîche."""
    ws = _fresh_sheet(wb, name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return ws


def _write_tableau_sheet(wb, students_df, summary):
    """Feuille matricielle : stats par classe (texte) puis élèves par classe.

    Sans sparklines (non gérées par openpyxl) ; les noms d'élèves restent
    désanonymisables par la macro qui parcourt toute la plage utilisée.
    """
    ws = _fresh_sheet(wb, "Tableau")
    students_df = students_df.sort_values(["classe", "student"])
    data = {cls: list(grp["student"]) for cls, grp in students_df.groupby("classe")}
    classes = list(data.keys())

    def stat(cl, key):
        if cl in summary.index and key in summary.columns:
            return int(summary.loc[cl, key])
        return 0

    ws.append([""] + classes)
    ws.append([""] + [f"F:{stat(c,'Filles')} G:{stat(c,'Garçons')}" for c in classes])
    ws.append([""] + [f"N1:{stat(c,'Niveau1')} N2:{stat(c,'Niveau2')} N3:{stat(c,'Niveau3')}" for c in classes])
    ws.append([""] + [f"C1:{stat(c,'Comp1')} C2:{stat(c,'Comp2')} C3:{stat(c,'Comp3')}" for c in classes])
    ws.append([""] + [f"POR:{stat(c,'POR')} LAT:{stat(c,'LAT')} TECH:{stat(c,'TECH')} CAT:{stat(c,'CAT')} PAP:{stat(c,'PAP')}" for c in classes])
    ws.append([])
    max_len = max((len(v) for v in data.values()), default=0)
    for i in range(max_len):
        ws.append([i + 1] + [data[c][i] if i < len(data[c]) else "" for c in classes])


def _reset_dictionary_sheet(wb):
    """Réinitialise `diccionari` (en-têtes seules) : aucune donnée résiduelle."""
    ws = _fresh_sheet(wb, "diccionari")
    ws["A1"] = "nom_real"
    ws["B1"] = "id_anonim"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
